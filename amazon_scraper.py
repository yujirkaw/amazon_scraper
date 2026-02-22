"""
Amazon.co.jp 商品ページスクレイパー

Excelファイル（Input/asin_list.xlsx）からASINリストを読み込み、
各商品ページから以下4項目を取得してExcelに出力する:
  B列: この商品について
  C列: メーカーによる説明
  D列: 商品情報
  E列: 商品の説明
"""

import os
import re
import time
import random
import requests
from bs4 import BeautifulSoup
from openpyxl import load_workbook
from openpyxl.styles import Alignment


BASE_DIR = os.path.dirname(os.path.abspath(__file__))
INPUT_DIR = os.path.join(BASE_DIR, "Input")
OUTPUT_DIR = os.path.join(BASE_DIR, "Output")
INPUT_FILE = os.path.join(INPUT_DIR, "asin_list.xlsx")
OUTPUT_FILE = os.path.join(OUTPUT_DIR, "asin_results.xlsx")

AMAZON_BASE_URL = "https://www.amazon.co.jp/dp/{asin}/?th=1"

HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/120.0.0.0 Safari/537.36"
    ),
    "Accept": (
        "text/html,application/xhtml+xml,application/xml;"
        "q=0.9,image/avif,image/webp,*/*;q=0.8"
    ),
    "Accept-Language": "ja,en-US;q=0.7,en;q=0.3",
    "Accept-Encoding": "gzip, deflate, br",
    "Connection": "keep-alive",
    "Upgrade-Insecure-Requests": "1",
}


def fetch_page(asin: str, max_retries: int = 3) -> BeautifulSoup | None:
    """Amazon商品ページを取得してBeautifulSoupオブジェクトを返す"""
    url = AMAZON_BASE_URL.format(asin=asin)
    session = requests.Session()
    session.headers.update(HEADERS)

    for attempt in range(max_retries):
        try:
            response = session.get(url, timeout=30)
            if response.status_code == 200:
                return BeautifulSoup(response.text, "lxml")
            print(f"  [WARN] ASIN {asin}: HTTP {response.status_code} (試行 {attempt + 1}/{max_retries})")
        except requests.RequestException as e:
            print(f"  [ERROR] ASIN {asin}: {e} (試行 {attempt + 1}/{max_retries})")

        if attempt < max_retries - 1:
            wait = 2 ** (attempt + 1) + random.uniform(0, 1)
            time.sleep(wait)

    return None


def extract_about_this_item(soup: BeautifulSoup) -> str:
    """「この商品について」を取得 (feature-bullets)"""
    # パターン1: #feature-bullets
    section = soup.find("div", id="feature-bullets")
    if section:
        items = section.find_all("span", class_="a-list-item")
        bullets = []
        for item in items:
            text = item.get_text(strip=True)
            if text:
                bullets.append(text)
        if bullets:
            return "\n".join(f"・{b}" for b in bullets)

    # パターン2: #productFactsDesktopExpander
    section = soup.find("div", id="productFactsDesktopExpander")
    if section:
        items = section.find_all("li")
        bullets = []
        for item in items:
            text = item.get_text(strip=True)
            if text:
                bullets.append(text)
        if bullets:
            return "\n".join(f"・{b}" for b in bullets)

    return ""


def extract_manufacturer_description(soup: BeautifulSoup) -> str:
    """「メーカーによる説明」を取得 (A+ content / aplus)"""
    section = soup.find("div", id="aplus")
    if not section:
        section = soup.find("div", id="aplus_feature_div")
    if not section:
        return ""

    texts = []
    # テキスト要素を取得
    for tag in section.find_all(["p", "h1", "h2", "h3", "h4", "h5", "span", "td"]):
        text = tag.get_text(strip=True)
        if text and text not in texts:
            texts.append(text)

    return "\n".join(texts) if texts else ""


def extract_product_information(soup: BeautifulSoup) -> str:
    """「商品情報」を取得 (Technical details / product details)"""
    rows = []

    # パターン1: #productDetails_techSpec_section_1 テーブル
    table = soup.find("table", id="productDetails_techSpec_section_1")
    if table:
        for tr in table.find_all("tr"):
            th = tr.find("th")
            td = tr.find("td")
            if th and td:
                key = th.get_text(strip=True)
                val = td.get_text(strip=True)
                rows.append(f"{key}: {val}")

    # パターン2: #productDetails_detailBullets_sections1 テーブル
    table2 = soup.find("table", id="productDetails_detailBullets_sections1")
    if table2:
        for tr in table2.find_all("tr"):
            th = tr.find("th")
            td = tr.find("td")
            if th and td:
                key = th.get_text(strip=True)
                val = td.get_text(strip=True)
                rows.append(f"{key}: {val}")

    # パターン3: #detailBullets_feature_div
    if not rows:
        detail_bullets = soup.find("div", id="detailBullets_feature_div")
        if detail_bullets:
            for li in detail_bullets.find_all("li"):
                spans = li.find_all("span", class_="a-text-bold")
                for span in spans:
                    key = span.get_text(strip=True).rstrip(":\u200f\u200e ")
                    sibling = span.find_next_sibling("span")
                    if sibling:
                        val = sibling.get_text(strip=True)
                        rows.append(f"{key}: {val}")

    # パターン4: #prodDetails 内の全テーブル
    if not rows:
        prod_details = soup.find("div", id="prodDetails")
        if prod_details:
            for tr in prod_details.find_all("tr"):
                th = tr.find("th")
                td = tr.find("td")
                if th and td:
                    key = th.get_text(strip=True)
                    val = td.get_text(strip=True)
                    rows.append(f"{key}: {val}")

    return "\n".join(rows) if rows else ""


def extract_product_description(soup: BeautifulSoup) -> str:
    """「商品の説明」を取得 (#productDescription)"""
    section = soup.find("div", id="productDescription")
    if section:
        # <p> タグのテキストを取得
        paragraphs = section.find_all("p")
        if paragraphs:
            texts = [p.get_text(strip=True) for p in paragraphs if p.get_text(strip=True)]
            if texts:
                return "\n".join(texts)
        # <p> がない場合はdiv全体のテキスト
        text = section.get_text(strip=True)
        # "商品の説明" ヘッダー部分を除去
        text = re.sub(r"^商品の説明\s*", "", text)
        return text

    return ""


def scrape_asin(asin: str) -> dict:
    """1つのASINから4項目を取得"""
    print(f"  取得中: {asin} ...")
    soup = fetch_page(asin)
    if soup is None:
        print(f"  [ERROR] ASIN {asin}: ページ取得失敗")
        return {
            "この商品について": "取得失敗",
            "メーカーによる説明": "取得失敗",
            "商品情報": "取得失敗",
            "商品の説明": "取得失敗",
        }

    result = {
        "この商品について": extract_about_this_item(soup),
        "メーカーによる説明": extract_manufacturer_description(soup),
        "商品情報": extract_product_information(soup),
        "商品の説明": extract_product_description(soup),
    }

    found = sum(1 for v in result.values() if v)
    print(f"  完了: {asin} ({found}/4 項目取得)")
    return result


def read_asin_list(filepath: str) -> list[str]:
    """ExcelファイルからASINリストを読み込む"""
    wb = load_workbook(filepath, read_only=True)
    ws = wb.active
    asins = []
    for row in ws.iter_rows(min_row=2, max_col=1, values_only=True):
        val = row[0]
        if val is not None:
            asin = str(val).strip()
            if asin:
                asins.append(asin)
    wb.close()
    return asins


def write_results(input_filepath: str, output_filepath: str, results: dict[str, dict]):
    """結果をExcelファイルに書き込む"""
    wb = load_workbook(input_filepath)
    ws = wb.active

    # ヘッダー行にB〜E列のヘッダーを追加
    headers = ["この商品について", "メーカーによる説明", "商品情報", "商品の説明"]
    for col_idx, header in enumerate(headers, start=2):  # B=2, C=3, D=4, E=5
        ws.cell(row=1, column=col_idx, value=header)

    # 各ASINの結果を書き込む
    for row_idx in range(2, ws.max_row + 1):
        asin = ws.cell(row=row_idx, column=1).value
        if asin is None:
            continue
        asin = str(asin).strip()
        if asin not in results:
            continue

        data = results[asin]
        for col_idx, key in enumerate(headers, start=2):
            cell = ws.cell(row=row_idx, column=col_idx, value=data.get(key, ""))
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    # 列幅の調整
    ws.column_dimensions["A"].width = 15
    ws.column_dimensions["B"].width = 50
    ws.column_dimensions["C"].width = 50
    ws.column_dimensions["D"].width = 50
    ws.column_dimensions["E"].width = 50

    wb.save(output_filepath)
    wb.close()


def main():
    # Input/Outputディレクトリの確認
    if not os.path.isdir(INPUT_DIR):
        print(f"[ERROR] Inputフォルダが見つかりません: {INPUT_DIR}")
        print("  Inputフォルダを作成し、asin_list.xlsx を配置してください。")
        return

    if not os.path.isfile(INPUT_FILE):
        print(f"[ERROR] 入力ファイルが見つかりません: {INPUT_FILE}")
        print("  Input/asin_list.xlsx を作成してください。")
        print("  A列1行目: 'ASIN' (ヘッダー)")
        print("  A列2行目以降: ASIN コード")
        return

    os.makedirs(OUTPUT_DIR, exist_ok=True)

    # ASINリスト読み込み
    asins = read_asin_list(INPUT_FILE)
    if not asins:
        print("[ERROR] ASINが見つかりません。Excelファイルを確認してください。")
        return

    print(f"取得対象ASIN数: {len(asins)}")
    print("-" * 50)

    # スクレイピング実行
    results = {}
    for i, asin in enumerate(asins):
        results[asin] = scrape_asin(asin)
        # リクエスト間隔を空ける（レート制限対策）
        if i < len(asins) - 1:
            wait = random.uniform(2, 5)
            time.sleep(wait)

    # 結果をExcelに書き込み
    print("-" * 50)
    print(f"結果を書き込み中: {OUTPUT_FILE}")
    write_results(INPUT_FILE, OUTPUT_FILE, results)
    print(f"完了! 出力ファイル: {OUTPUT_FILE}")


if __name__ == "__main__":
    main()
