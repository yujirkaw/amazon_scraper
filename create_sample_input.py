"""サンプル入力ファイルを作成するスクリプト"""
from openpyxl import Workbook

wb = Workbook()
ws = wb.active
ws.title = "ASIN List"

ws.cell(row=1, column=1, value="ASIN")
ws.cell(row=2, column=1, value="B09DX1R4RQ")

ws.column_dimensions["A"].width = 15

wb.save("Input/asin_list.xlsx")
print("サンプル入力ファイルを作成しました: Input/asin_list.xlsx")
